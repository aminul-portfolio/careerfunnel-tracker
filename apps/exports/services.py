from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.applications.models import JobApplication
from apps.daily_log.models import DailyLog
from apps.interviews.models import InterviewPrep
from apps.notes.models import Note
from apps.weekly_review.models import WeeklyReview

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
HEADER_TEXT_FONT = Font(bold=True)


def excel_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat(sep=" ") if hasattr(value, "hour") else value.isoformat()
    return value


def style_header_row(worksheet, row_number: int = 1) -> None:
    for cell in worksheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 4, 45)


def freeze_header(worksheet) -> None:
    worksheet.freeze_panes = "A2"


def workbook_to_response_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def create_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def add_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Read Me First")
    rows = [
        ["CareerFunnel Tracker Export"],
        ["Purpose"],
        ["This workbook exports your job-search tracker data for review and backup."],
        [],
        ["Main Diagnostic Logic"],
        ["High applications + low responses = likely CV / targeting issue"],
        ["Responses + no screening calls = likely messaging / role-fit issue"],
        ["Screening calls + no interviews = likely screening-call performance issue"],
        ["Interviews + no offers = likely interview / technical evidence issue"],
        ["Low daily activity = consistency / execution issue"],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = TITLE_FONT
    sheet["A2"].font = HEADER_TEXT_FONT
    sheet["A5"].font = HEADER_TEXT_FONT
    autosize_columns(sheet)


def add_applications_sheet(workbook: Workbook, user) -> None:
    sheet = workbook.create_sheet("Applications")
    headers = [
        "Company",
        "Job Title",
        "Job URL",
        "Location",
        "Work Type",
        "Salary Range",
        "Source",
        "Role Fit",
        "Date Applied",
        "Status",
        "Response Date",
        "Days to Response",
        "CV Version",
        "Cover Letter Version",
        "Contact Name",
        "Contact Email",
        "Notes",
    ]
    sheet.append(headers)
    for application in JobApplication.objects.filter(user=user):
        sheet.append(
            [
                excel_value(v)
                for v in [
                    application.company_name,
                    application.job_title,
                    application.job_url,
                    application.location,
                    application.get_work_type_display(),
                    application.salary_range,
                    application.get_source_display(),
                    application.get_role_fit_display(),
                    application.date_applied,
                    application.get_status_display(),
                    application.response_date,
                    application.days_to_response,
                    application.cv_version,
                    application.cover_letter_version,
                    application.contact_name,
                    application.contact_email,
                    application.notes,
                ]
            ]
        )
    style_header_row(sheet)
    freeze_header(sheet)
    autosize_columns(sheet)


def add_daily_logs_sheet(workbook: Workbook, user) -> None:
    sheet = workbook.create_sheet("Daily Logs")
    headers = [
        "Date",
        "Target Applications",
        "Actual Applications",
        "Variance",
        "Activity Signal",
        "Cover Letters Drafted",
        "Responses Received",
        "Calls Received",
        "Hours Spent",
        "Energy Level",
        "Notes",
    ]
    sheet.append(headers)
    for log in DailyLog.objects.filter(user=user):
        sheet.append(
            [
                excel_value(v)
                for v in [
                    log.log_date,
                    log.target_applications,
                    log.actual_applications,
                    log.variance,
                    log.activity_signal,
                    log.cover_letters_drafted,
                    log.responses_received,
                    log.calls_received,
                    log.hours_spent,
                    log.energy_level,
                    log.notes,
                ]
            ]
        )
    style_header_row(sheet)
    freeze_header(sheet)
    autosize_columns(sheet)


def add_weekly_reviews_sheet(workbook: Workbook, user) -> None:
    sheet = workbook.create_sheet("Weekly Reviews")
    headers = [
        "Week Starting",
        "Week Ending",
        "Target Applications",
        "Actual Applications",
        "Application Variance",
        "Responses Received",
        "Screening Calls",
        "Technical Screens",
        "Interviews",
        "Offers",
        "Rejections",
        "Diagnosis",
        "Mood",
        "Response Rate",
        "Screening Rate",
        "Interview Rate",
        "Offer Rate",
        "What Worked",
        "What Blocked",
        "Lessons Learned",
        "Change Next Week",
    ]
    sheet.append(headers)
    for review in WeeklyReview.objects.filter(user=user):
        sheet.append(
            [
                excel_value(v)
                for v in [
                    review.week_starting,
                    review.week_ending,
                    review.target_applications,
                    review.actual_applications,
                    review.application_variance,
                    review.responses_received,
                    review.screening_calls,
                    review.technical_screens,
                    review.interviews,
                    review.offers,
                    review.rejections,
                    review.get_diagnosis_display(),
                    review.get_mood_display(),
                    review.response_rate,
                    review.screening_rate,
                    review.interview_rate,
                    review.offer_rate,
                    review.what_worked,
                    review.what_blocked,
                    review.lessons_learned,
                    review.change_next_week,
                ]
            ]
        )
    style_header_row(sheet)
    freeze_header(sheet)
    autosize_columns(sheet)


def add_notes_sheet(workbook: Workbook, user) -> None:
    sheet = workbook.create_sheet("Notes Decisions")
    headers = [
        "Type",
        "Title",
        "Content",
        "Tags",
        "Decision Date",
        "Important",
        "Created At",
        "Updated At",
    ]
    sheet.append(headers)
    for note in Note.objects.filter(user=user):
        sheet.append(
            [
                excel_value(v)
                for v in [
                    note.get_note_type_display(),
                    note.title,
                    note.content,
                    note.tags,
                    note.decision_date,
                    "Yes" if note.is_important else "No",
                    note.created_at,
                    note.updated_at,
                ]
            ]
        )
    style_header_row(sheet)
    freeze_header(sheet)
    autosize_columns(sheet)



def add_interviews_sheet(workbook: Workbook, user) -> None:
    sheet = workbook.create_sheet("Interview Prep")
    headers = [
        "Company",
        "Role",
        "Interview Date",
        "Stage",
        "Outcome",
        "Readiness Score",
        "Expected Topics",
        "Projects to Mention",
        "Questions to Prepare",
        "Reflection",
    ]
    sheet.append(headers)
    for interview in InterviewPrep.objects.filter(user=user):
        sheet.append([
            interview.application.company_name,
            interview.application.job_title,
            interview.interview_date,
            interview.get_stage_display(),
            interview.get_outcome_display(),
            interview.readiness_score,
            interview.expected_topics,
            interview.projects_to_mention,
            interview.questions_to_prepare,
            interview.reflection,
        ])
    style_header_row(sheet)
    freeze_header(sheet)
    autosize_columns(sheet)


def build_interviews_workbook(user) -> bytes:
    workbook = create_workbook()
    add_interviews_sheet(workbook, user)
    return workbook_to_response_bytes(workbook)


def build_applications_workbook(user) -> bytes:
    workbook = create_workbook()
    add_applications_sheet(workbook, user)
    return workbook_to_response_bytes(workbook)


def build_daily_logs_workbook(user) -> bytes:
    workbook = create_workbook()
    add_daily_logs_sheet(workbook, user)
    return workbook_to_response_bytes(workbook)


def build_weekly_reviews_workbook(user) -> bytes:
    workbook = create_workbook()
    add_weekly_reviews_sheet(workbook, user)
    return workbook_to_response_bytes(workbook)


def build_notes_workbook(user) -> bytes:
    workbook = create_workbook()
    add_notes_sheet(workbook, user)
    return workbook_to_response_bytes(workbook)


def build_full_tracker_workbook(user) -> bytes:
    workbook = create_workbook()
    add_readme_sheet(workbook)
    add_applications_sheet(workbook, user)
    add_daily_logs_sheet(workbook, user)
    add_weekly_reviews_sheet(workbook, user)
    add_notes_sheet(workbook, user)
    add_interviews_sheet(workbook, user)
    return workbook_to_response_bytes(workbook)
